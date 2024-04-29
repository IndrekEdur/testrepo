from django.shortcuts import render, redirect, reverse
from django.core.paginator import PageNotAnInteger, EmptyPage
from django.contrib.auth.decorators import login_required
from datetime import datetime
from django.contrib.auth.models import User
from django.views.generic import ListView, UpdateView, DeleteView, CreateView, DetailView
from django.urls import reverse_lazy
from django.http import JsonResponse
import json
import random

from . models import Product, Project, Comment, Hours
from . forms import ProductForm, CommentForm, HoursForm, GeneralHoursForm, ProjectForm

import smtplib
from email.message import EmailMessage

from .decorators import allowed_users

from .models import Question, Quiz
from django.core.paginator import Paginator

from products.tasks import send_facebook_message

from .models import Article

from django.http import HttpResponseRedirect
import os
import logging

from django.http import HttpResponse
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from fuzzywuzzy import fuzz, process
import re
from django.contrib import messages
from collections import defaultdict

logger = logging.getLogger(__name__)
logger.debug("Test log message from my_views")


logger = logging.getLogger(__name__)
logger.debug("Test log message")

def send_messenger_message (request):
    logger.debug("This is a debug message in my_view")
    if request.method == 'POST':
        # Define your recipient_id and message_text (or get them from the request)
        recipient_id = "429716328007568"  # Replace with actual recipient ID
        message_text = "Hello!"  # Replace with actual message text
        page_access_token = os.environ.get('FACEBOOK_PAGE_ACCESS_TOKEN')
        send_facebook_message()

        # Redirect to a new URL, or inform the user of success
        return HttpResponseRedirect(reverse('showUsers'))  # Replace with the name of a success URL

        # If not a POST request, you can redirect or show a form
    return HttpResponseRedirect(reverse('showProjects'))  # Replace with a home or form URL


import tempfile  # Import the tempfile module

def calculate_match_score(query_tokens, article_tokens):
    score = 0
    total_weight = 0

    for token in query_tokens:
        if token.isdigit():  # Higher weight for numeric tokens
            token_weight = 2
            if token in article_tokens:
                score += 100 * token_weight  # High score for exact numeric match
        else:
            token_weight = 1
            # Using a partial ratio to allow for partial matches within longer strings
            match = process.extractOne(token, article_tokens, scorer=fuzz.partial_ratio)
            if match:
                score += match[1] * token_weight
        total_weight += token_weight

    return score / total_weight if total_weight else 0

def extract_matched_parts(query, best_match_name):
    # This is a simplified version, it can be expanded as needed
    return best_match_name, "Unmatched part not implemented yet"

def tokenize(text):
    """
    Extracts complex numerical patterns and words as tokens.
    Adjusted regex to handle '4x0,8+0,8' as a single token.
    """
    # This pattern looks for sequences of digits possibly separated by 'x', ',', or '+'
    # and may include other digits, ensuring such patterns are captured as whole tokens.
    pattern = r'\b\d+[xX,]*\d*(?:\+\d+)?\b|\b[a-zA-Z]+\b'
    return re.findall(pattern, text.lower())

def get_article_matches(query_words, articles):
    """ Filters articles based on exact or near-exact presence of query words. """
    matches = defaultdict(list)
    for word in query_words:
        for article in articles:
            # Consider an article a match if it contains the query word
            if word in tokenize(article):
                matches[word].append(article)
    return matches

def process_and_match_articles(request):
    if request.method == 'POST' and request.FILES['myfile']:
        excel_file = request.FILES['myfile']
        file_extension = excel_file.name.split('.')[-1]
        if file_extension not in ['xlsx', 'xlsm', 'xltx', 'xltm']:
            return HttpResponse("Unsupported file format", status=400)

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            for chunk in excel_file.chunks():
                temp_file.write(chunk)
            temp_file_path = temp_file.name

        df = pd.read_excel(temp_file_path)
        combined_filter = df.apply(lambda x: pd.to_numeric(x, errors='coerce').notna()).any(axis=1)
        filtered_df = df[combined_filter]

        output = Workbook()
        sheet = output.active
        for r in dataframe_to_rows(filtered_df, index=False, header=True):
            sheet.append(r)

        article_names = list(Article.objects.all().values_list('name', flat=True))
        sheet.insert_cols(1)
        sheet.insert_cols(2)
        sheet.insert_cols(3)
        sheet.insert_cols(4)

        for idx, row in enumerate(list(sheet.rows)[1:], start=2):
            full_row_text = ' '.join(str(cell.value) for cell in row[5:] if cell.value)
            query_words = tokenize(full_row_text)
            matches = get_article_matches(query_words, article_names)
            potential_matches = {article for match_list in matches.values() for article in match_list}

            best_match = process.extractOne(full_row_text, list(potential_matches), scorer=fuzz.token_sort_ratio) if potential_matches else (None, 0)
            best_match_name, match_score = best_match

            sheet.cell(row=idx, column=1, value=best_match_name if best_match_name else "No match found")
            sheet.cell(row=idx, column=2, value=match_score if best_match_name else 0)
            sheet.cell(row=idx, column=3, value=', '.join(potential_matches))
            sheet.cell(row=idx, column=4, value='; '.join(query_words))  # Use semicolon to separate query words

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="enhanced_filtered_data.xlsx"'
        output.save(response)
        os.unlink(temp_file_path)

        return response

    return render(request, 'products/upload.html')

def upload_articles(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            data = pd.read_excel(file)
            articles = []
            for index, row in data.iterrows():
                # Normalize text fields to lowercase
                article_code_lower = str(row['Code']).lower().strip()
                name_lower = str(row['Name']).lower().strip()
                description_lower = str(row.get('Description', '')).lower().strip()

                # Clean the price field to remove non-numeric characters and convert to float
                price_text = str(row['Price'])
                cleaned_price = re.sub(r'[^\d.]+', '', price_text)
                price = float(cleaned_price) if cleaned_price else 0

                # Update or create the article in the database
                article, created = Article.objects.update_or_create(
                    article_code=article_code_lower,
                    defaults={
                        'name': name_lower,
                        'price': price,
                        'description': description_lower
                    }
                )
                articles.append(article)
            messages.success(request, f"{len(articles)} articles uploaded successfully")
            return redirect('upload-articles')
    else:
        form = UploadFileForm()
    return render(request, 'products/upload_articles.html', {'form': form})

def smart_split(input_string):
    # Pattern to identify numbers followed by letters with optional commas
    pattern = r'(\d+[\.,]?\d*\s*[a-zA-Z]+)'
    # Split the string by spaces and commas not followed by numbers (to keep decimal numbers together)
    parts = re.split(r',\s*(?!\d)', input_string)
    tokens = []
    for part in parts:
        # Further split each part if it matches the pattern
        sub_parts = re.findall(pattern, part)
        if sub_parts:
            tokens.extend(sub_parts)
        else:
            tokens.append(part)
    return [token.strip().replace(',', '') for token in tokens if token.strip()]  # Remove any stray commas



def extract_matched_parts(name_to_match, best_match):
    name_words = set(name_to_match.split())
    match_words = set(best_match.split())
    matched_parts = name_words & match_words
    unmatched_parts = name_words - match_words

    matched_string = ' '.join(sorted(matched_parts))
    unmatched_string = ' '.join(sorted(unmatched_parts))
    return matched_string, unmatched_string

import os
import re
from django.shortcuts import render
from .forms import DocumentForm
from .models import Document
from openpyxl import Workbook
import openpyxl
from django.shortcuts import render, get_object_or_404

def medical_quiz_data_modification(request):
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            # Load the uploaded file
            doc = Document.objects.latest('id')
            wb = openpyxl.load_workbook(doc.file.path, data_only=True)
            ws = wb.active

            # Create new Workbooks for questions and answers
            questions_wb = Workbook()
            answers_wb = Workbook()
            ws_questions = questions_wb.active
            ws_answers = answers_wb.active

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                if row[0].value is not None:  # Ensure the row is not empty
                    content = str(row[0].value).strip()
                    # Determine if it's a question or an answer
                    if content[0].isdigit() or not (content[1:3] == ') ' and content[0].isalpha()):
                        # Extract question number from the beginning of the question text
                        match = re.match(r'^(\d+)', content)
                        if match:
                            question_num = int(match.group(1))
                        else:
                            question_num = None  # Fallback if no number found
                        # Remove the numbers and potentially other non-question text if necessary
                        question_text = re.sub(r'^\d+\.\s*', '', content)
                        ws_questions.append([question_text])  # Append processed question text
                    else:
                        # Extract the correctness indicator and remove the label
                        correctness = int(row[1].value) if row[1].value else 0
                        label_removed = content[3:].strip() if content[2] == ' ' else content[2:].strip()
                        ws_answers.append([label_removed, correctness, question_num])  # Append answer with indicator and question number

            # Define the paths for saving the processed files
            save_path = os.path.expanduser('~\\Downloads')
            questions_filename = os.path.join(save_path, 'processed_questions.xlsx')
            answers_filename = os.path.join(save_path, 'processed_answers.xlsx')
            questions_wb.save(questions_filename)
            answers_wb.save(answers_filename)

            return render(request, 'products/input_medical_questions.html', {
                'form': form,
                'success': True,
                'questions_file': questions_filename,
                'answers_file': answers_filename
            })
    else:
        form = DocumentForm()
    return render(request, 'products/input_medical_questions.html', {'form': form, 'success': False})






class ProductListView(ListView):
    model = Product
    template_name = 'products/product_list.html'
    context_object_name = 'products'

def quiz(request, group=1):
    questions_per_group = 20
    start = (group - 1) * questions_per_group
    end = start + questions_per_group

    questions = Question.objects.all()[start:end]
    total_groups = (Question.objects.count() - 1) // questions_per_group + 1

    return render(request, 'products/quiz.html', {
        'questions': questions,
        'current_group': group,
        'total_groups': range(1, total_groups + 1)
    })

def quiz_selection(request):
    quizzes = Quiz.objects.all()
    return render(request, 'products/quiz_selection.html', {'quizzes': quizzes})

def random_quiz(request, quiz_id):
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    total_questions = 20
    questions = list(quiz.questions.all())
    random_questions = random.sample(questions, min(total_questions, len(questions)))
    return render(request, 'products/random_quiz.html', {'questions': random_questions, 'quiz_id': quiz_id})


def submit_random_quiz(request, quiz_id):
    if request.method == 'POST':
        question_ids = request.POST.getlist('question_ids')
        questions = Question.objects.filter(id__in=question_ids).prefetch_related('answers')
        score = 0
        results = []

        for question in questions:
            user_answers_ids = request.POST.getlist(f'question_{question.id}')
            all_answers = question.answers.all()

            answer_feedback = []

            for answer in all_answers:
                user_selected = str(answer.id) in user_answers_ids
                is_correct = answer.is_correct
                if user_selected and is_correct:
                    score += 1

                answer_feedback.append({
                    'text': answer.answer_text,
                    'user_selected': user_selected,
                    'is_correct': is_correct
                })

            results.append({
                'question': question.question_text,
                'answers': answer_feedback
            })

        context = {
            'results': results,
            'score': score,
            'total': len(questions)
        }
        return render(request, 'products/quiz_results.html', {'quiz_id': quiz_id})


def submit_quiz(request):
    if request.method == 'POST':
        group_num = int(request.POST.get('group_num', 1))  # Retrieve the group number
        questions_per_group = 20
        start = (group_num - 1) * questions_per_group
        end = start + questions_per_group

        questions = Question.objects.all()[start:end].prefetch_related('answers')
        score = 0
        results = []

        for question in questions:
            user_answers_ids = request.POST.getlist(f'question_{question.id}')
            all_answers = question.answers.all()

            answer_feedback = []

            for answer in all_answers:
                user_selected = str(answer.id) in user_answers_ids
                is_correct = answer.is_correct
                if user_selected and is_correct:
                    score += 1

                answer_feedback.append({
                    'text': answer.answer_text,
                    'user_selected': user_selected,
                    'is_correct': is_correct
                })

            results.append({
                'question': question.question_text,
                'answers': answer_feedback
            })

        context = {
            'results': results,
            'score': score,
            'total': len(questions)  # Number of questions in the current group
        }
        return render(request, 'products/quiz_results.html', context)



def quiz_results(request):
    # Retrieve results from session or database
    results = request.session.get('quiz_results', [])
    score = request.session.get('quiz_score', 0)
    total = request.session.get('quiz_total', 0)

    return render(request, 'products/quiz_results.html', {
        'results': results,
        'score': score,
        'total': total
    })



@login_required(login_url='accounts/login.html')
def ShowAllProjects(request):
    projects = Project.objects.order_by('-project_number')
    context = {'projects': projects}
    return render(request, 'products/project_list.html', context)

@login_required(login_url='accounts/login.html')
def ShowAllUsers(request):
    users = User.objects.all
    context = {'users': users}
    return render(request, 'products/users_list.html', context)

@login_required(login_url='accounts/login.html')
def ShowAllProducts(request):

    project = request.GET.get('project')


    if project == None:
        products = Product.objects.order_by('-created_at').filter(is_completed=False)
        dropboxlink = 'none'
        page_num = request.GET.get("page")
        paginator = Paginator(products, 8)
        try:
            products = paginator.page(page_num)
        except PageNotAnInteger:
            products = paginator.page(1)
        except EmptyPage:
            products = paginator.page(paginator.num_pages)
    else:
        products = Product.objects.order_by('-created_at').filter(Project__name=project, is_completed=False)
        dropboxlink = 'none'
    projects = Project.objects.all()

    context = {
        'products': products,
        'projects': projects,
        'dropboxlink': dropboxlink
     }

    return render(request, 'products/showProducts.html', context)

def updateItem(request):
    data = json.loads(request.body)
    productId = data['productId']
    productId = data['action']
    user = request.user
    product = Product.objects.get(id=productId)
    # create an order
    # create cart in session
    cartItems = request.session.get('cartItems', 0)
    request.session['cartItems'] = cartItems + 1

    return JsonResponse({'message': 'Item was added'})


class ProductCreateView(CreateView):
    model = Product
    Product.objects.order_by().filter()
    template_name = 'products/product_create.html'
    fields = '__all__'
    success_url = reverse_lazy('product_list')

@login_required(login_url='showProducts')
def addProduct(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('showProducts')
    else:
        form = ProductForm()
    context = {
        "form": form
     }
    return render(request, 'products/addProduct.html', context)

@login_required(login_url='showProducts')
def addHours(request, pk):

    product = Product.objects.get(id=pk)
    project = product.Project
    form = HoursForm(instance=product)
    if request.method == 'POST':
        form = HoursForm(request.POST, instance=product)
        if form.is_valid():
            owner = request.user.id
            quantity = form.cleaned_data['quantity']
            date = form.cleaned_data['date']
            c = Hours(project=product.Project, product=product, date=date, owner=request.user, quantity=quantity, inserted_at=datetime.now())
            c.save()
            return redirect('showProducts')
    else:
        form = HoursForm()

    return render(request, 'products/addHours.html', {'form': form })

@login_required(login_url='showProducts')
def addGeneralHours(request):

    if request.method == 'POST':
        form = GeneralHoursForm(request.POST)
        if form.is_valid():
            owner = form.cleaned_data['owner']
            quantity = form.cleaned_data['quantity']
            date = form.cleaned_data['date']
            product = form.cleaned_data['product']
            project = form.cleaned_data['project']

            c = Hours(product=product, project=project, date=date, owner=owner, quantity=quantity, inserted_at=datetime.now())
            c.save()
            return redirect('showProducts')
    else:
        form = GeneralHoursForm()

    return render(request, 'products/addHoursGeneral.html', {'form': form })

class ProjectCreateView(CreateView):
    model = Project
    template_name = 'products/project_create.html'
    fields = '__all__'
    success_url = reverse_lazy('showProducts')

class ProductDetailView(DetailView):
    model = Product
    template_name = 'products/product_detail.html'
    context_object_name = 'product'

@login_required(login_url='showProducts')
def ProjectGeneralView(request, pk):
    hours_sum_aggregate = Hours.objects.filter(project_id=pk).aggregate(Sum('quantity'))
    try: hours_sum = float(hours_sum_aggregate['quantity__sum'])
    except: hours_sum = 0
    project = Project.objects.get(id=pk)
    context = { "project": project, "sum": hours_sum, "pk": pk}

    return render(request, 'products/project_general.html', context)

@login_required(login_url='showProducts')
@allowed_users(allowed_roles=['Project_manager'])
def ProjectDetailView(request, pk):
    hours_sum_aggregate = Hours.objects.filter(project_id=pk).aggregate(Sum('quantity'))
    try: hours_sum = float(hours_sum_aggregate['quantity__sum'])
    except: hours_sum = 0
    project = Project.objects.get(id=pk)
    hours = Hours.objects.order_by('-date')
    context = { "project": project, "hours": hours, "product": Product.objects.all, "sum": hours_sum, "pk": pk}

    return render(request, 'products/project_detail.html', context)


@login_required(login_url='showProducts')
def productDetail(request, pk):
    eachProduct = Product.objects.get(id=pk)

    num_comments = Comment.objects.filter(product=eachProduct).count()

    context = {
        'eachProduct': eachProduct,
        'num_comments': num_comments,
     }

    return render(request, 'products/productDetail.html', context)

class ProductUpdateView(UpdateView):
    model = Product
    template_name = 'products/product_update.html'
    context_object_name = 'product'
    fields = '__all__'
    success_url = reverse_lazy('product_list')

class HoursUpdateView(UpdateView):
    model = Hours
    template_name = 'products/hours_update.html'
    context_object_name = 'hours'
    fields = '__all__'
    success_url = reverse_lazy('showProjects')

@login_required(login_url='showProducts')
def updateProduct(request, pk):
    product = Product.objects.get(id=pk)
    form = ProductForm(instance=product)

    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            sendNotification(request, pk, product.author.email, " : has been updated! ")
            sendNotification(request, pk, product.performer.email, " : has been updated! ")
            return redirect('showProducts')

    context = {
        "form": form,
    }

    return render(request, 'products/updateProduct.html', context)

@login_required(login_url='showProducts')
@allowed_users(allowed_roles=['Project_manager'])
def updateProject(request, pk):
    project = Project.objects.get(id=pk)
    form = ProjectForm(instance=project)

    if request.method == 'POST':
        form = ProjectForm(request.POST, request.FILES, instance=project)
        if form.is_valid():
            form.save()

            return redirect('showProjects')

    context = { "form": form}

    return render(request, 'products/updateProject.html', context)

@login_required(login_url='sendNotification')
def sendNotification(request, pk, email, content):
    product = Product.objects.get(id=pk)

    msg = EmailMessage()
    msg.set_content( product.Project.name + " : " + product.name + "  >>>  " + content + "  Here is the link for details: http://indreke.pythonanywhere.com/products/product/"+str(pk)+"/" + "   Description: " + product.description)
    msg['subject'] = product.Project.name + " : " + product.name
    msg['to'] = email

    user = "erlin.virtuaalne.assistent@gmail.com"
    msg['from'] = user
    password = "ipjykvcufysqigsc"

    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()
    server.login(user, password)
    server.send_message(msg)

    server.quit()

    return redirect('showProducts')

class ProductDeleteView(DeleteView):
    model = Product
    template_name = 'products/product_delete.html'
    context_object_name = 'product'
    success_url = reverse_lazy('product_list')

class HoursDeleteView(DeleteView):
    model = Hours
    template_name = 'products/hours_delete.html'
    context_object_name = 'hours'
    success_url = reverse_lazy('showProjects')


@login_required(login_url='showProducts')
@allowed_users(allowed_roles=['Project_manager'])
def deleteProduct(request, pk):
    product = Product.objects.get(id=pk)
    sendNotification(request, pk, product.author.email, " : has been deleted! ")
    sendNotification(request, pk, product.performer.email, " : has been deleted! ")
    product.delete()
    return redirect('showProducts')


@login_required(login_url='showProducts')
def searchBar(request):
    if request.method == 'GET':
        query = request.GET.get('query')
        if query:
            products = Product.objects.filter(hours__icontains=query)
            return render(request, 'products/searchbar.html', {'products': products})
        else:
            print("No information to show")
            return render(request, 'products/searchbar.html', {})


def add_comment(request, pk):
    eachProduct = Product.objects.get(id=pk)

    form = CommentForm(instance=eachProduct)

    if request.method == 'POST':
        form = CommentForm(request.POST, instance=eachProduct)
        product = Product.objects.get(id=pk)
        if form.is_valid():
            name = request.user.username
            body = form.cleaned_data['comment_body']
            c = Comment(product=eachProduct, commenter_name=name, comment_body=body, date_added=datetime.now())
            c.save()
            sendNotification(request, pk, product.author.email, " : has been commented! ")
            sendNotification(request, pk, product.performer.email, " : has been commented! ")
            return redirect('showProducts')
        else:
            print('form is invalid')
    else:
        form = CommentForm()

    context = {
        'form': form
    }
    return render(request, 'products/add_comment.html', context)

def delete_comment(request, pk):
    comment = Comment.objects.filter(product=pk).last()
    product_id = comment.product.id
    comment.delete()
    return redirect(reverse('product', args=[product_id]))

from django.db.models import Sum


@login_required(login_url='showProducts')
def WorkingHoursListView(request, pk):
    hours_sum_aggregate = Hours.objects.filter(owner_id=pk).aggregate(Sum('quantity'))
    hours_sum = float(hours_sum_aggregate['quantity__sum'])
    hours = Hours.objects.filter(owner_id=pk).order_by('-date')
    months=["12", "11", "10", "9", "8", "7", "6", "5", "4", "3", "2"]
    context = { "months": months, "hours": hours, "product": Product.objects.all, "sum": hours_sum}

    return render(request, 'products/hours_list.html', context)

@login_required(login_url='showProducts')
def WorkingHoursListByProjectsView(request, pk):
    hours_sum_aggregate = Hours.objects.filter(owner_id=pk).aggregate(Sum('quantity'))
    hours_sum = float(hours_sum_aggregate['quantity__sum'])


    user_hours = Hours.objects.filter(owner_id=pk)
    user_projects_list = list(Hours.objects.filter(owner_id=pk))

    distinct_user_project_ids = list(Hours.objects.all().values_list('project_id', flat=True).distinct())

    user_projects_and_hours = list(Hours.objects.filter(owner_id=pk).values('project_id', 'quantity'))

    project_hours_sum_aggregate = Hours.objects.filter(owner_id=pk, project_id=4).aggregate(Sum('quantity'))

    projects = Project.objects.all

    q = Hours.objects.annotate(project_hours_sum=Sum('quantity'))
    customers = Hours.objects.annotate(orders_count=Sum('quantity'))

    context = { "x": user_projects_and_hours, "projects": projects, "projects_list": distinct_user_project_ids, "hours": user_hours, "product": Product.objects.all, "sum": hours_sum}

    return render(request, 'products/hours_list_by_project.html', context)

@login_required(login_url='showProducts')
@allowed_users(allowed_roles=['Project_manager'])
def UserHoursListView(request, pk):
    hours_sum_aggregate = Hours.objects.filter(owner_id=pk).aggregate(Sum('quantity'))
    hours_sum = (hours_sum_aggregate['quantity__sum'])
    hours = Hours.objects.filter(owner_id=pk).order_by('-date')
    months = ["12", "11", "10", "9", "8", "7", "6", "5", "4", "3", "2", "1"]
    context = { "months": months, "hours": hours, "product": Product.objects.all, "sum": hours_sum}

    return render(request, 'products/hours_list_of_user.html', context)


@login_required(login_url='showProducts')
def WorkingHoursSum(request, pk,project):
    hours_sum_aggregate = Hours.objects.filter(owner_id=pk, project_id=project).aggregate(Sum('quantity'))
    hours_sum = float(hours_sum_aggregate['quantity__sum'])

    context = {"sum": hours_sum}

    return render(request, context)